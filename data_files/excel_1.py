from pathlib import Path
from openpyxl import Workbook, load_workbook

# Create a new workbook and write data
name = "Test_1"
wb = Workbook()
sheet = wb.active
sheet.title = name
sheet["A1"] = "Name"
sheet["B1"] = "Age"
sheet["A2"] = "Priyanka"
sheet["B2"] = 23
sheet.append(["Alice", 25])
sheet.append(["Bob", 30])
sheet.cell(row=5, column=1, value="Charlie")
sheet.cell(row=5, column=2, value=35)
path = Path(__file__).parent.joinpath("sample.xlsx")
wb.save(path)
print("Excel file created and data written.")

# Load the workbook and read data
wb = load_workbook(path)
sheet = wb[name]
for row in sheet.iter_rows(values_only=True):
    print(row)

# Update a cell
sheet["B2"] = 26  # Update Alice's age
path = Path(__file__).parent.joinpath("sample_updated.xlsx")
wb.save(path)
print("Excel file updated.")
